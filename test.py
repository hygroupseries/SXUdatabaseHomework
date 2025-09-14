import sqlite3
import pandas as pd
from openpyxl import load_workbook  # 用于获取工作表名称

def excel_to_sqlite(excel_file, db_file, table_name=None, sheet_name=None):
    """
    将Excel数据导入SQLite3数据库
    
    参数:
        excel_file (str): Excel文件路径
        db_file (str): SQLite数据库文件路径
        table_name (str): 目标表名(可选，默认为Excel工作表名)
        sheet_name (str): 要导入的工作表名(可选，默认为第一个工作表)
    """
    try:
        # 连接到SQLite数据库(如果不存在会自动创建)
        conn = sqlite3.connect(db_file)
        cursor = conn.cursor()
        
        # 如果没有指定工作表名，获取第一个工作表名
        if sheet_name is None:
            wb = load_workbook(excel_file, read_only=True)
            sheet_name = wb.sheetnames[0]
            wb.close()
        
        # 使用pandas读取Excel文件
        print(f"正在读取Excel文件: {excel_file} 的工作表: {sheet_name}")
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        
        # 确定表名
        if table_name is None:
            # 使用工作表名作为表名，替换空格和特殊字符
            table_name = sheet_name.replace(" ", "_").replace("-", "_").lower()
        
        # 将数据写入SQLite数据库
        print(f"正在将数据导入到SQLite表: {table_name}")
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        
        # 提交事务并关闭连接
        conn.commit()
        print("数据导入成功!")
        
        # 显示导入的表结构
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        print("\n表结构:")
        for col in columns:
            print(f"{col[1]} ({col[2]})")
            
    except Exception as e:
        print(f"发生错误: {e}")
        conn.rollback()
    finally:
        if conn:
            conn.close()

if __name__ == "__main__":
    # 用户输入
    excel_file = input("请输入Excel文件路径: ").strip()
    db_file = input("请输入SQLite数据库文件路径(如不存在将创建): ").strip()
    sheet_name = input("请输入要导入的工作表名(留空使用第一个工作表): ").strip() or None
    table_name = input("请输入目标表名(留空使用工作表名): ").strip() or None
    
    # 执行导入
    excel_to_sqlite(excel_file, db_file, table_name, sheet_name)